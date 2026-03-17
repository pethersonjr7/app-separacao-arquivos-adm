import os
import re
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter

# ==============================
# CONFIGURE OS CAMINHOS AQUI
# ==============================

pasta_entrada = r"C:\TESTE\ORIGEN"
pasta_saida = r"C:\TESTE\DEST"

# Encontra automaticamente o primeiro arquivo Excel na pasta de entrada
arquivos_na_pasta = [f for f in os.listdir(pasta_entrada) if f.endswith(('.xlsx', '.xlsm'))]
if not arquivos_na_pasta:
    print(f"ERRO: Nenhum arquivo Excel encontrado em {pasta_entrada}")
    exit()

nome_arquivo = arquivos_na_pasta[0]
print(f"Arquivo selecionado para processamento: {nome_arquivo}")

# ==============================
# FUNÇÃO PARA LIMPAR NOME DO ARQUIVO
# ==============================

def limpar_nome_arquivo(nome):
    return re.sub(r'[\\/*?:"<>|]', "", str(nome)).strip()

# ==============================
# PROCESSAMENTO
# ==============================

if __name__ == "__main__":
    caminho_arquivo = os.path.join(pasta_entrada, nome_arquivo)

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    COL_REPRESENTANTE = 7  # Coluna G

    cabecalho = list(ws.iter_rows(min_row=1, max_row=1))

    dados_por_representante = {}

    for row in ws.iter_rows(min_row=2):
        representante = row[COL_REPRESENTANTE - 1].value
        if representante:
            dados_por_representante.setdefault(representante, []).append(row)

    for representante, linhas in dados_por_representante.items():
        
        nome_limpo = limpar_nome_arquivo(representante)
        novo_wb = Workbook()
        novo_ws = novo_wb.active
        novo_ws.title = "Dados"

        # Copiar largura das colunas
        for col in ws.column_dimensions:
            novo_ws.column_dimensions[col].width = ws.column_dimensions[col].width

        # Copiar cabeçalho com estilo
        for row in cabecalho:
            for cell in row:
                nova_cell = novo_ws.cell(row=1, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    nova_cell.font = copy(cell.font)
                    nova_cell.border = copy(cell.border)
                    nova_cell.fill = copy(cell.fill)
                    nova_cell.number_format = copy(cell.number_format)
                    nova_cell.alignment = copy(cell.alignment)

        # Copiar dados com estilo
        for i, row in enumerate(linhas, start=2):
            for cell in row:
                nova_cell = novo_ws.cell(row=i, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    nova_cell.font = copy(cell.font)
                    nova_cell.border = copy(cell.border)
                    nova_cell.fill = copy(cell.fill)
                    nova_cell.number_format = copy(cell.number_format)
                    nova_cell.alignment = copy(cell.alignment)

        # ==============================
        # ATIVAR FILTRO NA PRIMEIRA LINHA
        # ==============================
        ultima_coluna = novo_ws.max_column
        ultima_linha = novo_ws.max_row
        
        letra_ultima_coluna = get_column_letter(ultima_coluna)
        novo_ws.auto_filter.ref = f"A1:{letra_ultima_coluna}{ultima_linha}"

        caminho_saida = os.path.join(pasta_saida, f"{nome_limpo}.xlsx")
        novo_wb.save(caminho_saida)

        print(f"Arquivo criado: {caminho_saida}")

    print("Processo finalizado com sucesso!")
