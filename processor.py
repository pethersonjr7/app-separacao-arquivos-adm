import os
import re
import io
import zipfile
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def limpar_nome_arquivo(nome):
    nome_limpo = re.sub(r'[\\/*?:"<>|]', "", str(nome)).strip()
    return nome_limpo if nome_limpo else "Sem_Nome"

def process_excel(input_file_stream, col_index=7):
    """
    Processes an Excel file stream and returns a dictionary of file contents and a log of events.
    Returns:
        {
            "files": { "filename.xlsx": bytes, ... },
            "logs": ["log message 1", ...]
        }
    """
    logs = []
    logs.append("Iniciando carregamento da planilha...")
    
    try:
        # data_only=False to preserve formulas if any, or True if formulas are breaking things.
        # Original script didn't specify, so we use default (False).
        wb = load_workbook(input_file_stream)
        ws = wb.active
        logs.append(f"Planilha carregada: {ws.title}")
    except Exception as e:
        logs.append(f"ERRO ao carregar planilha: {str(e)}")
        raise e

    if ws.max_column < col_index:
        msg = f"ERRO: A planilha possui apenas {ws.max_column} colunas, mas a coluna {col_index} foi solicitada."
        logs.append(msg)
        raise ValueError(msg)

    logs.append(f"Analisando coluna {get_column_letter(col_index)} para separação...")
    
    cabecalho = list(ws.iter_rows(min_row=1, max_row=1))
    
    dados_por_representante = {}
    rows_processed = 0
    for row in ws.iter_rows(min_row=2):
        # Garantir que a linha tem colunas suficientes
        if len(row) < col_index:
            continue
            
        valor_separador = row[col_index - 1].value
        if valor_separador:
            valor_str = str(valor_separador).strip()
            dados_por_representante.setdefault(valor_str, []).append(row)
            rows_processed += 1

    logs.append(f"Processadas {rows_processed} linhas de dados.")
    logs.append(f"Encontrados {len(dados_por_representante)} representantes únicos.")

    output_files = {}

    # Cache para estilos para evitar copy() redundantes
    style_cache = {
        'font': {},
        'fill': {},
        'border': {},
        'alignment': {},
        'number_format': {}
    }

    def get_cached_style(cell, style_type):
        original_style = getattr(cell, style_type)
        if original_style is None:
            return None
        
        # Usar o hash do objeto de estilo como chave
        try:
            style_hash = hash(original_style)
        except TypeError:
            # Se não for hashable, copia normalmente
            return copy(original_style)
            
        if style_hash not in style_cache[style_type]:
            style_cache[style_type][style_hash] = copy(original_style)
        return style_cache[style_type][style_hash]

    for representante, linhas in dados_por_representante.items():
        base_nome = limpar_nome_arquivo(representante)
        
        # Garantir nome único se houver colisão após limpeza
        nome_arquivo = f"{base_nome}.xlsx"
        counter = 1
        while nome_arquivo in output_files:
            nome_arquivo = f"{base_nome}_{counter}.xlsx"
            counter += 1

        logs.append(f"Gerando arquivo para: {representante} ({len(linhas)} linhas)...")
        
        try:
            novo_wb = Workbook()
            novo_ws = novo_wb.active
            novo_ws.title = "Dados"

            # Copiar largura das colunas de forma mais robusta
            for key in ws.column_dimensions:
                if key in ws.column_dimensions:
                    novo_ws.column_dimensions[key].width = ws.column_dimensions[key].width

            # Copiar cabeçalho com estilo
            for row in cabecalho:
                for cell in row:
                    try:
                        col_idx = cell.column if isinstance(cell.column, int) else cell.col_idx
                        nova_cell = novo_ws.cell(row=1, column=col_idx, value=cell.value)
                        if cell.has_style:
                            nova_cell.font = get_cached_style(cell, 'font')
                            nova_cell.border = get_cached_style(cell, 'border')
                            nova_cell.fill = get_cached_style(cell, 'fill')
                            nova_cell.number_format = cell.number_format # Number format is just a string
                            nova_cell.alignment = get_cached_style(cell, 'alignment')
                    except Exception as e_style:
                        print(f"Erro ao copiar estilo do cabeçalho: {e_style}")

            # Copiar dados sem estilo para máxima performance de CPU e Memória
            for row in linhas:
                values = [cell.value for cell in row]
                novo_ws.append(values)

            # Ativar filtro na primeira linha
            ultima_coluna = novo_ws.max_column
            ultima_linha = novo_ws.max_row
            if ultima_coluna > 0 and ultima_linha > 0:
                letra_ultima_coluna = get_column_letter(ultima_coluna)
                novo_ws.auto_filter.ref = f"A1:{letra_ultima_coluna}{ultima_linha}"

            # Save to a bytes buffer
            excel_buffer = io.BytesIO()
            novo_wb.save(excel_buffer)
            output_files[nome_arquivo] = excel_buffer.getvalue()
        except Exception as e_proc:
            logs.append(f"ERRO ao gerar arquivo para {representante}: {str(e_proc)}")
            print(f"Erro no processamento do representante {representante}: {e_proc}")

    logs.append("Processamento finalizado!")
    return {
        "files": output_files,
        "logs": logs
    }
