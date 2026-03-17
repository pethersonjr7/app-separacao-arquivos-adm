import sys
import os
import traceback
from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
import uuid
import zipfile
import io
import shutil
from processor import process_excel

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Configurações básicas
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64MB limit
TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_outputs')

# Garantir que a pasta temporária existe
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/headers', methods=['POST'])
def get_headers():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Nenhum arquivo selecionado"}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        headers = []
        # Read the first row for headers
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    col_idx = cell.column if isinstance(cell.column, int) else cell.col_idx
                    headers.append({
                        "index": col_idx,
                        "name": str(cell.value),
                        "letter": get_column_letter(col_idx)
                    })
        return jsonify({"headers": headers})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Nenhum arquivo selecionado"}), 400

    col_index = request.form.get('column_index', default=7, type=int)

    if not file.filename.endswith(('.xlsx', '.xlsm')):
        return jsonify({"error": "Apenas arquivos Excel (.xlsx, .xlsm) são permitidos"}), 400

    try:
        # Process the file
        result = process_excel(file, col_index=col_index)
        
        # Create a unique ID for this request
        request_id = str(uuid.uuid4())
        request_path = os.path.join(TEMP_DIR, request_id)
        os.makedirs(request_path)

        # Save files to disk temporarily for individual downloads
        file_list = []
        for filename, content in result['files'].items():
            file_path = os.path.join(request_path, filename)
            with open(file_path, 'wb') as f:
                f.write(content)
            file_list.append(filename)

        # Create ZIP
        zip_path = os.path.join(request_path, 'planilhas_separadas.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename, content in result['files'].items():
                zipf.writestr(filename, content)

        return jsonify({
            "request_id": request_id,
            "original_filename": os.path.splitext(file.filename)[0],
            "files": file_list,
            "logs": result['logs']
        })
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/download/<request_id>/<filename>')
def download_file(request_id, filename):
    return send_from_directory(os.path.join(TEMP_DIR, request_id), filename, as_attachment=True)

@app.route('/download_all/<request_id>')
def download_zip(request_id):
    return send_from_directory(os.path.join(TEMP_DIR, request_id), 'planilhas_separadas.zip', as_attachment=True)

if __name__ == '__main__':
    # Use a porta definida pelo ambiente (Cloud Run) ou 3000 para local
    port = int(os.environ.get('PORT', 3000))
    app.run(debug=True, host='0.0.0.0', port=port)
